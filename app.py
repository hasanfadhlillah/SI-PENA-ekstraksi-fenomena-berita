# File: app.py
import re
import streamlit as st
import pandas as pd
import os
import time
import json
import io
import altair as alt
from datetime import datetime, timedelta
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from radar.query_expander import dapatkan_keywords, _load_keywords, _save_keywords
from radar.database import (
    inisialisasi_database,
    ambil_semua_status_kategori,
    ambil_artikel_valid,
    tandai_artikel_diekstrak,
    tandai_artikel_ditolak,
    simpan_hasil_ekstraksi,
    get_connection,
    reset_total_database,
)
from radar.pipeline import scan_kategori, batch_scan_semua_kategori, _hitung_triwulan
from radar.config import DEFAULT_MIN_SKOR
from radar.backup import (
    buat_backup_keywords_bytes,
    buat_backup_database_bytes,
    pulihkan_keywords_dari_upload,
    auto_restore_dari_hf_dataset,
    force_backup_ke_hf_dataset,
)
from scraper import scrape_berita
from ai_engine import ekstrak_fenomena_ai
from radar.logger_config import get_logger
logger = get_logger("app")

auto_restore_dari_hf_dataset()

# ─── KONFIGURASI HALAMAN ───────────────────────────────────────────────────────
st.set_page_config(
    page_title="SI-PENA | BPS Kota Magelang",
    page_icon="✒️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS Custom
st.markdown("""
<style>
.kartu-artikel {
    border: 1px solid rgba(130, 130, 130, 0.2);
    border-radius: 10px;
    padding: 16px;
    margin-bottom: 12px;
    background-color: rgba(130, 130, 130, 0.05);
}
.kartu-artikel a { color: #4a6cf7 !important; word-break: break-all; text-decoration: none; }
.kartu-artikel a:hover { text-decoration: underline; }
.badge-skor-hijau {
    background-color: rgba(40, 167, 69, 0.2); color: #28a745;
    padding: 4px 10px; border-radius: 20px; font-weight: bold; font-size: 13px;
}
.badge-skor-kuning {
    background-color: rgba(255, 193, 7, 0.2); color: #d39e00;
    padding: 4px 10px; border-radius: 20px; font-weight: bold; font-size: 13px;
}
.alasan-box {
    background-color: rgba(74, 108, 247, 0.1);
    border-left: 4px solid #4a6cf7;
    padding: 10px 14px; border-radius: 4px; font-size: 14px; margin-top: 8px;
}
/* ─── TENTANG PAGE ─── */
.hero-banner {
    background: linear-gradient(135deg, #003366 0%, #0055a5 60%, #1a73e8 100%);
    border-radius: 16px;
    padding: 48px 40px;
    margin-bottom: 32px;
    text-align: center;
    color: white;
}
.hero-banner h1 { font-size: 3.2rem; font-weight: 900; margin: 0 0 8px 0; letter-spacing: 1px; }
.hero-banner .tagline { font-size: 1.15rem; font-weight: 600; color: #a8d4ff; margin-bottom: 12px; }
.hero-banner .desc { font-size: 0.95rem; color: #cce4ff; max-width: 680px; margin: 0 auto; line-height: 1.7; }
.fitur-card {
    border: 1px solid rgba(74, 108, 247, 0.25);
    border-radius: 12px;
    padding: 22px 20px;
    height: 100%;
    background: rgba(74, 108, 247, 0.04);
    transition: transform 0.2s;
}
.fitur-card:hover { transform: translateY(-2px); }
.fitur-card .icon { font-size: 2rem; margin-bottom: 10px; }
.fitur-card h4 { font-size: 1rem; font-weight: 700; color: #003366; margin: 0 0 8px 0; }
.fitur-card p { font-size: 0.88rem; color: #555; line-height: 1.6; margin: 0; }
.alur-step {
    display: flex; align-items: flex-start; gap: 16px;
    padding: 16px; border-radius: 10px;
    background: rgba(0, 51, 102, 0.04);
    border-left: 4px solid #003366;
    margin-bottom: 12px;
}
.alur-step .nomor {
    background: #003366; color: white;
    border-radius: 50%; width: 32px; height: 32px;
    display: flex; align-items: center; justify-content: center;
    font-weight: 900; font-size: 0.9rem; flex-shrink: 0;
}
.alur-step .konten h5 { margin: 0 0 4px 0; font-size: 0.95rem; font-weight: 700; color: #003366; }
.alur-step .konten p { margin: 0; font-size: 0.85rem; color: #555; line-height: 1.5; }
.tech-chip {
    display: inline-block;
    padding: 5px 14px;
    border-radius: 20px;
    font-size: 0.82rem;
    font-weight: 600;
    margin: 4px;
}
.skor-row-green { background: rgba(40,167,69,0.12); border-left: 4px solid #28a745; padding: 6px 12px; border-radius: 4px; margin-bottom: 5px; font-size: 0.88rem; }
.skor-row-yellow { background: rgba(255,193,7,0.12); border-left: 4px solid #ffc107; padding: 6px 12px; border-radius: 4px; margin-bottom: 5px; font-size: 0.88rem; }
.skor-row-red { background: rgba(220,53,69,0.10); border-left: 4px solid #dc3545; padding: 6px 12px; border-radius: 4px; margin-bottom: 5px; font-size: 0.88rem; }
.geo-lolos { background: rgba(40,167,69,0.10); border-left: 3px solid #28a745; padding: 6px 12px; border-radius: 4px; margin-bottom: 4px; font-size: 0.87rem; }
.geo-tolak { background: rgba(220,53,69,0.10); border-left: 3px solid #dc3545; padding: 6px 12px; border-radius: 4px; margin-bottom: 4px; font-size: 0.87rem; }
</style>
""", unsafe_allow_html=True)

inisialisasi_database()
load_dotenv()

# ─── MENGAMBIL API KEYS ───────────────────────────────────────────────────────
KEYS = {
    "groq"    : os.environ.get("GROQ_API_KEYS", os.environ.get("GROQ_API_KEY", "")),
    "gemini"  : os.environ.get("GEMINI_API_KEYS", os.environ.get("GEMINI_API_KEY", "")),
    "cerebras": os.environ.get("CEREBRAS_API_KEYS", os.environ.get("CEREBRAS_API_KEY", "")),
    "mistral" : os.environ.get("MISTRAL_API_KEYS", os.environ.get("MISTRAL_API_KEY", "")),
}

def _hitung_kunci(raw_keys: str) -> int:
    return len([k for k in raw_keys.split(",") if k.strip()])

# ─── Standar nama file download ───────────────────────────────────────────────
def _nama_file(tipe: str, fmt: str) -> str:
    """
    Standar penamaan file download.
    tipe: 'EkstraksiFenomena' | 'RiwayatRadar' | 'SemuaEkstraksi'
    fmt : 'xlsx' | 'csv' | 'json'
    """
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    tgl = datetime.now().strftime("%Y%m%d")
    if tipe == "EkstraksiFenomena":
        return f"SIPENA_{tipe}_{ts}.{fmt}"
    return f"SIPENA_{tipe}_{tgl}.{fmt}"

DAFTAR_KATEGORI = [
    "Tanaman Pangan", "Tanaman Hortikultura Semusim", "Perkebunan Semusim",
    "Tanaman Hortikultura Tahunan", "Perkebunan Tahunan", "Peternakan",
    "Jasa Pertanian dan Perburuan", "Kehutanan dan Penebangan Kayu", "Perikanan",
    "Pertambangan Minyak dan Gas Bumi", "Pertambangan Batubara dan Lignit",
    "Pertambangan Bijih Logam", "Pertambangan dan Penggalian Lainnya",
    "Industri Makanan dan Minuman", "Pengolahan Tembakau",
    "Industri Tekstil dan Pakaian Jadi", "Industri Kulit, Barang dari Kulit dan Alas Kaki",
    "Industri Kayu, Barang dari Kayu dan Gabus", "Industri Kertas dan Percetakan",
    "Industri Kimia, Farmasi dan Obat Tradisional", "Industri Karet, Barang dari Karet dan Plastik",
    "Industri Barang Galian bukan Logam", "Industri Logam Dasar",
    "Industri Barang dari Logam, Komputer, Elektronik", "Industri Alat Angkutan", "Industri Furnitur",
    "Industri Pengolahan Lainnya, Jasa Reparasi, Pemasangan Mesin dan Peralatan",
    "Ketenagalistrikan", "Pengadaan Gas dan Produksi Es", "Pengadaan Air", "Konstruksi",
    "Perdagangan Mobil, Sepeda Motor dan Reparasinya", "Perdagangan Besar dan Eceran",
    "Angkutan Rel", "Angkutan Darat", "Angkutan Laut", "Angkutan Udara",
    "Pergudangan dan Jasa Penunjang Angkutan", "Penyediaan Akomodasi", "Penyediaan Makan Minum",
    "Informasi dan Komunikasi", "Jasa Perantara Keuangan", "Asuransi dan Dana Pensiun",
    "Jasa Keuangan Lainnya", "Real Estate", "Jasa Perusahaan",
    "Administrasi Pemerintahan dan Jaminan Sosial", "Jasa Pendidikan",
    "Jasa Kesehatan dan Kegiatan Sosial", "Jasa Lainnya", "PRODUK DOMESTIK BRUTO"
]

# Validasi otomatis sinkronisasi DAFTAR_KATEGORI vs keywords.json
def _validasi_sinkronisasi_kategori():
    try:
        kw_data = _load_keywords()
        set_dropdown = set(DAFTAR_KATEGORI)
        set_keywords = set(kw_data.keys())

        hilang_dari_dropdown = set_keywords - set_dropdown
        hilang_dari_keywords = set_dropdown - set_keywords

        if hilang_dari_dropdown or hilang_dari_keywords:
            with st.sidebar:
                with st.expander("⚠️ Peringatan: Sinkronisasi Kategori", expanded=False):
                    if hilang_dari_dropdown:
                        st.warning(
                            f"**{len(hilang_dari_dropdown)} kategori** ada di `keywords.json` "
                            f"tapi TIDAK ADA di dropdown Radar (tidak bisa dipilih user):"
                        )
                        for k in sorted(hilang_dari_dropdown):
                            st.caption(f"• {k}")
                    if hilang_dari_keywords:
                        st.warning(
                            f"**{len(hilang_dari_keywords)} kategori** ada di dropdown Radar "
                            f"tapi TIDAK ADA di `keywords.json` (akan otomatis pakai AI fallback "
                            f"untuk generate keyword saat dipilih):"
                        )
                        for k in sorted(hilang_dari_keywords):
                            st.caption(f"• {k}")
    except Exception as e:
        logger.error(f"[Validasi Kategori] Gagal cek sinkronisasi: {e}")

# ─── Helper: konversi nilai AI ke string aman ─────────────────────────────────
def _ke_str(nilai) -> str:
    if isinstance(nilai, (dict, list)):
        return json.dumps(nilai, ensure_ascii=False, indent=2)
    return str(nilai) if nilai else ""

# ─── Helper Excel: style dasar ────────────────────────────────────────────────
def _style_excel_dasar():
    """Return kumpulan style yang dipakai oleh semua fungsi Excel."""
    header_fill    = PatternFill("solid", fgColor="003366")
    subheader_fill = PatternFill("solid", fgColor="DDEEFF")
    warna_ganjil   = PatternFill("solid", fgColor="F7FAFF")
    warna_genap    = PatternFill("solid", fgColor="FFFFFF")
    border_tipis   = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap_align   = Alignment(wrap_text=True, vertical="top")
    return (header_fill, subheader_fill, warna_ganjil, warna_genap,
            border_tipis, center_align, wrap_align)

# ─── Excel: Satu Hasil Ekstraksi (Tab 2) ─────────────────────────────────────
def _buat_excel_ekstraksi(json_final: dict) -> bytes:
    wb  = Workbook()
    ws  = wb.active
    ws.title = "Ekstraksi Fenomena"
    (header_fill, subheader_fill, warna_ganjil, warna_genap,
     border_tipis, center_align, wrap_align) = _style_excel_dasar()

    ws.merge_cells("A1:C1")
    ws["A1"] = "FORMULIR EKSTRAKSI FENOMENA EKONOMI — BPS KOTA MAGELANG"
    ws["A1"].font      = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    ws["A1"].fill      = header_fill
    ws["A1"].alignment = center_align
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:C2")
    ws["A2"] = (f"Diekstrak pada: {json_final.get('_waktu_ekstraksi', '')}  |  "
                f"Model AI: {json_final.get('_model_digunakan', '-')}  |  "
                f"URL: {json_final.get('_url_sumber', '')}")
    ws["A2"].font      = Font(name="Calibri", italic=True, size=9, color="555555")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 8

    subheader_font = Font(name="Calibri", bold=True, color="003366", size=10)
    for col_idx, h in enumerate(["No.", "Variabel BPS", "Hasil Ekstraksi (Bisa Diedit)"], 1):
        cell           = ws.cell(row=4, column=col_idx, value=h)
        cell.font      = subheader_font
        cell.fill      = subheader_fill
        cell.border    = border_tipis
        cell.alignment = center_align
    ws.row_dimensions[4].height = 22

    LABEL_MAP = [
        ("tema_topik",            "1. Tema / Topik"),
        ("judul_dan_tanggal",     "2. Judul & Tanggal Terbit"),
        ("sumber_dan_link",       "3. Sumber & Link Media"),
        ("ringkasan_fenomena",    "4. Ringkasan Fenomena"),
        ("data_angka",            "5. Data Angka Kuantitatif"),
        ("kutipan_tokoh",         "6. Kutipan Tokoh / Narasumber"),
        ("lokasi_spesifik",       "7. Lokasi Spesifik"),
        ("intervensi_pemerintah", "8. Intervensi Pemerintah"),
        ("periode_kejadian",      "9. Periode Kejadian"),
        ("kata_kunci",            "10. Kata Kunci / Hashtag"),
        ("sentimen_dampak",       "11. Sentimen Dampak"),
        ("kategori_perbandingan", "12. Kategori Perbandingan"),
    ]
    isi_font = Font(name="Calibri", size=10)

    for i, (key, label) in enumerate(LABEL_MAP):
        row        = i + 5
        nilai      = json_final.get(key, "")
        if isinstance(nilai, (dict, list)):
            nilai = json.dumps(nilai, ensure_ascii=False, indent=2)
        fill = warna_ganjil if i % 2 == 0 else warna_genap

        c_no           = ws.cell(row=row, column=1, value=i+1)
        c_no.font      = Font(name="Calibri", size=10, bold=True)
        c_no.alignment = center_align
        c_no.border    = border_tipis
        c_no.fill      = fill

        c_var           = ws.cell(row=row, column=2, value=label)
        c_var.font      = Font(name="Calibri", size=10, bold=True, color="003366")
        c_var.alignment = Alignment(vertical="top", wrap_text=True)
        c_var.border    = border_tipis
        c_var.fill      = fill

        c_val           = ws.cell(row=row, column=3, value=str(nilai))
        c_val.font      = isi_font
        c_val.alignment = wrap_align
        c_val.border    = border_tipis
        c_val.fill      = fill

        ws.row_dimensions[row].height = (
            80 if key in ["ringkasan_fenomena", "kutipan_tokoh",
                          "data_angka", "intervensi_pemerintah"] else 25
        )

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 75
    ws.freeze_panes = "A5"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


# ─── Excel: Riwayat Radar (Tab 3) ─────────────────────────────────────────────
def _buat_excel_riwayat(df: pd.DataFrame) -> bytes:
    """Excel berformat profesional untuk tabel Riwayat Pencarian Radar."""
    wb  = Workbook()
    ws  = wb.active
    ws.title = "Riwayat Radar SI-PENA"
    (header_fill, subheader_fill, warna_ganjil, warna_genap,
     border_tipis, center_align, wrap_align) = _style_excel_dasar()

    n_col    = len(df.columns)
    last_col = get_column_letter(n_col)

    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = "RIWAYAT PENCARIAN BERITA RADAR — BPS KOTA MAGELANG"
    ws["A1"].font      = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    ws["A1"].fill      = header_fill
    ws["A1"].alignment = center_align
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = f"Dicetak pada: {datetime.now().strftime('%d %B %Y  %H:%M')} WIB"
    ws["A2"].font      = Font(name="Calibri", italic=True, size=9, color="555555")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 6

    subheader_font = Font(name="Calibri", bold=True, color="003366", size=10)
    for col_idx, col_name in enumerate(df.columns, 1):
        cell           = ws.cell(row=4, column=col_idx, value=col_name)
        cell.font      = subheader_font
        cell.fill      = subheader_fill
        cell.border    = border_tipis
        cell.alignment = center_align
    ws.row_dimensions[4].height = 22

    isi_font = Font(name="Calibri", size=10)
    for row_idx, row_data in enumerate(df.itertuples(index=False), 5):
        fill = warna_ganjil if (row_idx - 5) % 2 == 0 else warna_genap
        for col_idx, value in enumerate(row_data, 1):
            cell           = ws.cell(row=row_idx, column=col_idx,
                                     value=(str(value) if value is not None else ""))
            cell.font      = isi_font
            cell.border    = border_tipis
            cell.fill      = fill
            cell.alignment = wrap_align
        ws.row_dimensions[row_idx].height = 20

    # Lebar kolom per nama kolom
    lebar = {
        "Judul Berita": 50, "Kategori PDRB": 28, "Triwulan": 12,
        "Skor AI": 10, "Status": 16, "Ditemukan": 20, "Diekstrak": 20,
    }
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = lebar.get(col_name, 18)

    ws.freeze_panes = "A5"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


# ─── Excel: Semua Hasil Ekstraksi (Tab 3) ─────────────────────────────────────
def _buat_excel_semua_ekstraksi(df: pd.DataFrame) -> bytes:
    """Excel berformat profesional untuk tabel Semua Hasil Ekstraksi Fenomena."""
    wb  = Workbook()
    ws  = wb.active
    ws.title = "Semua Ekstraksi Fenomena"
    (header_fill, subheader_fill, warna_ganjil, warna_genap,
     border_tipis, center_align, wrap_align) = _style_excel_dasar()

    n_col    = len(df.columns)
    last_col = get_column_letter(n_col)

    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = "REKAP SEMUA HASIL EKSTRAKSI FENOMENA EKONOMI — BPS KOTA MAGELANG"
    ws["A1"].font      = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    ws["A1"].fill      = header_fill
    ws["A1"].alignment = center_align
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = (f"Total {len(df)} rekaman ekstraksi  |  "
                f"Dicetak: {datetime.now().strftime('%d %B %Y  %H:%M')} WIB")
    ws["A2"].font      = Font(name="Calibri", italic=True, size=9, color="555555")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 6

    subheader_font = Font(name="Calibri", bold=True, color="003366", size=10)
    for col_idx, col_name in enumerate(df.columns, 1):
        cell           = ws.cell(row=4, column=col_idx, value=col_name)
        cell.font      = subheader_font
        cell.fill      = subheader_fill
        cell.border    = border_tipis
        cell.alignment = center_align
    ws.row_dimensions[4].height = 22

    # Kolom mana saja yang butuh baris tinggi karena konten panjang
    KOLOM_PANJANG = {
        "Ringkasan Fenomena", "Kutipan Tokoh", "Data Angka", "Intervensi Pemerintah"
    }
    isi_font = Font(name="Calibri", size=10)
    for row_idx, row_data in enumerate(df.itertuples(index=False), 5):
        fill = warna_ganjil if (row_idx - 5) % 2 == 0 else warna_genap
        tinggi = 18
        for col_idx, (value, col_name) in enumerate(zip(row_data, df.columns), 1):
            cell           = ws.cell(row=row_idx, column=col_idx,
                                     value=(str(value) if value is not None else ""))
            cell.font      = isi_font
            cell.border    = border_tipis
            cell.fill      = fill
            cell.alignment = wrap_align
            if col_name in KOLOM_PANJANG:
                tinggi = max(tinggi, 60)
        ws.row_dimensions[row_idx].height = tinggi

    lebar = {
        "Waktu Ekstraksi": 18, "Tema/Topik": 22, "Judul & Tanggal": 32,
        "Sumber & Link": 36, "Ringkasan Fenomena": 50, "Data Angka": 30,
        "Kutipan Tokoh": 40, "Lokasi Spesifik": 22, "Intervensi Pemerintah": 35,
        "Periode Kejadian": 18, "Kata Kunci": 20, "Sentimen": 12,
        "Jenis Perbandingan": 18, "Model AI": 26,
    }
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = lebar.get(col_name, 20)

    ws.freeze_panes = "A5"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


# ─── SESSION STATE ────────────────────────────────────────────────────────────
for key, default in [
    ("target_url", ""),
    ("hasil_ekstraksi", None),
    ("ekstraksi_url_aktif", ""),
    ("json_final_siap", None),
    ("kategori_terpilih_antrean", "— Pilih Kategori —"),
    ("backup_keywords_bytes", None),
    ("backup_db_bytes", None),
]:
    if key not in st.session_state:
        st.session_state[key] = default

# ─── SIDEBAR ──────────────────────────────────────────────────────────────────
with st.sidebar:
    if os.path.exists("logo_bps_magelang.png"):
        st.image("logo_bps_magelang.png", width='stretch')

    st.markdown("""
    <div style="margin-top: -10px; margin-bottom: 20px;">
        <h1 style="font-size: 2.8rem; font-weight: 900; margin-bottom: 0px; line-height: 1.1;">✒️ SI-PENA</h1>
        <div style="font-size: 1.05rem; font-weight: 800; color: #4a6cf7; margin-top: 5px;">Sistem Informasi Pencari Berita & Ekstraksi Fenomena</div>
        <div style="font-size: 0.85rem; font-style: italic; color: #828282; margin-top: 4px;">"Menulis ulang fakta dari ribuan narasi berita"</div>
        <div style="font-size: 0.95rem; font-weight: 600; margin-top: 8px;">BPS Kota Magelang</div>
    </div>
    <hr style="margin-top: 5px; margin-bottom: 15px;">
    """, unsafe_allow_html=True)

    st.markdown("### 📅 Rentang Waktu Pencarian")
    default_end   = datetime.now()
    default_start = default_end - timedelta(days=90)
    tanggal_mulai   = st.date_input("Dari Tanggal",   default_start, help="Batas awal berita diterbitkan.", key="date_mulai")
    tanggal_selesai = st.date_input("Sampai Tanggal", default_end,   help="Batas akhir berita diterbitkan.", key="date_selesai")

    if tanggal_mulai > tanggal_selesai:
        st.error("Tanggal mulai tidak boleh setelah tanggal selesai!")

    try:
        triwulan_berjalan = _hitung_triwulan(tanggal_mulai.strftime("%Y-%m-%d"))
        st.success(f"📌 Periode: **{triwulan_berjalan}**")
    except ValueError as e:
        st.error(f"⚠️ {e}")
        st.stop()

    st.markdown("### 🎛️ Pengaturan AI Radar")
    min_skor    = st.slider("Skor Minimum Lolos", 1, 10, DEFAULT_MIN_SKOR,
                            help="Filter seberapa ketat AI menyeleksi berita.", key="slider_min_skor")
    paksa_ulang = st.toggle("🔄 Proses Ulang Artikel Lama", value=False,
                            help="Jika diaktifkan, Radar akan men-scan ulang berita yang pernah ditolak/gagal.",
                            key="toggle_paksa_ulang")
    scan_semua  = st.toggle("🌐 Scan Semua Level Wilayah (khusus Batch Scan)", value=False,
                            help="Toggle ini HANYA berlaku untuk mode '✨ SEMUA KATEGORI (BATCH SCAN)'. "
                                 "Nonaktif (default) = Batch Scan berhenti begitu tiap kategori menemukan "
                                 "minimal 3 artikel. Catatan: scan 1 kategori (bukan Batch Scan) SELALU memindai semua "
                                 "5 level wilayah secara penuh, tidak terpengaruh toggle ini.",
                            key="toggle_scan_semua")

    st.markdown("---")
    st.markdown("### 🚦 Status Pasukan AI (Pool)")
    st.caption("Aplikasi ini menggunakan sistem Load Balancing. AI akan otomatis berganti kunci jika terjadi limit.")
    for nama, key_val in [("Groq", KEYS["groq"]), ("Cerebras", KEYS["cerebras"]),
                           ("Gemini", KEYS["gemini"]), ("Mistral", KEYS["mistral"])]:
        jumlah = _hitung_kunci(key_val)
        status = f"🟢 {jumlah} Amunisi Siap" if jumlah > 0 else "🔴 Kosong"
        st.caption(f"**{nama}:** {status}")
    st.markdown("---")

    # Backup & Restore (jaring pengaman ephemeral storage HF Spaces)
    with st.expander("💾 Backup & Restore Data", expanded=False):
        st.caption(
            "⚠️ Aplikasi ini di-deploy di Hugging Face Spaces (tier gratis) yang "
            "TIDAK menyimpan data secara permanen — setiap kali Space di-restart "
            "atau tidur lalu bangun lagi, riwayat artikel & keyword BISA HILANG. "
            "Siapkan & unduh backup secara berkala, terutama sebelum jeda pemakaian lama."
        )
        if st.button("🔄 Siapkan File Backup Terbaru", key="btn_siapkan_backup"):
            st.session_state["backup_keywords_bytes"] = buat_backup_keywords_bytes()
            st.session_state["backup_db_bytes"] = buat_backup_database_bytes()
            st.toast("Backup siap diunduh di bawah!", icon="✅")

        if st.session_state["backup_keywords_bytes"] is not None:
            col_bk1, col_bk2 = st.columns(2)
            with col_bk1:
                st.download_button(
                    "⬇️ keywords.json",
                    data=st.session_state["backup_keywords_bytes"],
                    file_name=f"backup_keywords_{datetime.now().strftime('%Y%m%d_%H%M')}.json",
                    mime="application/json",
                    width='stretch',
                    key="dl_backup_keywords",
                )
            with col_bk2:
                st.download_button(
                    "⬇️ Database",
                    data=st.session_state["backup_db_bytes"],
                    file_name=f"backup_sifeno_tracker_{datetime.now().strftime('%Y%m%d_%H%M')}.db",
                    mime="application/octet-stream",
                    width='stretch',
                    key="dl_backup_db",
                )
        else:
            st.caption("Klik tombol di atas untuk menyiapkan file sebelum diunduh.")

        st.markdown("**🔄 Pulihkan keywords.json dari backup:**")
        file_restore = st.file_uploader(
            "Upload file backup keywords.json:", type=["json"],
            label_visibility="collapsed", key="restore_keywords_uploader"
        )
        if file_restore is not None:
            if st.button("♻️ Pulihkan Sekarang", key="btn_restore_keywords"):
                berhasil, pesan = pulihkan_keywords_dari_upload(file_restore.read())
                if berhasil:
                    st.success(pesan)
                    st.rerun()
                else:
                    st.error(pesan)

        _hf_token_ada = bool(os.environ.get("HF_TOKEN", ""))
        _hf_repo_ada  = bool(os.environ.get("HF_BACKUP_REPO_ID", ""))
        
        if _hf_token_ada and _hf_repo_ada:
            st.success(
                "🟢 Auto-backup & Auto-restore ke HF Dataset AKTIF — data otomatis "
                "tersimpan tiap kategori selesai discan, dan otomatis dipulihkan "
                "kalau Space baru saja restart."
            )
        else:
            st.info(
                "🔴 Auto-backup ke HF Dataset belum aktif. Set secret `HF_TOKEN` dan "
                "`HF_BACKUP_REPO_ID` di Settings Space untuk mengaktifkan backup otomatis gratis."
            )
        
        st.markdown("---")
        st.markdown("**🗑️ Reset Total**")
        st.caption(
            "⚠️ **PERINGATAN:** Ini akan menghapus SELURUH riwayat artikel, "
            "status kategori, dan hasil ekstraksi dari database (Mulai dari 0) — TIDAK BISA DIBATALKAN. "
            "Jika auto-backup HF Dataset aktif, backup lama juga langsung ditimpa "
            "dengan versi kosong supaya tidak ke-restore lagi saat Space restart."
        )
        konfirmasi_reset = st.text_input(
            "Ketik **HAPUS SEMUA** (persis, huruf besar) untuk mengaktifkan tombol reset:",
            key="input_konfirmasi_reset"
        )
        if st.button("🗑️ RESET TOTAL",
                     type="secondary", width='stretch',
                     disabled=(konfirmasi_reset != "HAPUS SEMUA"),
                     key="btn_reset_total"):
            reset_total_database()
            berhasil_force_backup = force_backup_ke_hf_dataset()
            st.session_state.kategori_terpilih_antrean = "— Pilih Kategori —"
            if berhasil_force_backup:
                st.success("✅ Reset total berhasil! Database kosong & backup HF Dataset sudah ditimpa dengan versi kosong.")
            else:
                st.success("✅ Reset total berhasil! Database sekarang kosong.")
                st.info("ℹ️ Auto-backup HF Dataset tidak aktif/gagal — tidak ada backup lama yang perlu ditimpa.")
            time.sleep(2)
            st.rerun()
        
    st.markdown("---")
    st.caption("v1.0 | Made with ❤️ for BPS Kota Magelang")

# ─── MAIN TABS ───────────────────────────────────────────────────────────────
TAB_LABELS = [
    "📡 RADAR BERITA",
    "📝 EKSTRAKTOR FENOMENA",
    "🗄️ HISTORY BERITA",
    "📈 DASHBOARD ANALISIS",
    "⚙️ KELOLA KEYWORD",
    "ℹ️ TENTANG SI-PENA",
]

st.markdown("""
<style>
div[data-testid="stRadio"] > div[role="radiogroup"] {
    flex-direction: row !important;
    flex-wrap: wrap;
    gap: 2px;
    border-bottom: 2px solid rgba(130,130,130,0.25);
    margin-bottom: 1.2rem;
    padding-bottom: 0;
}
div[data-testid="stRadio"] > div[role="radiogroup"] > label {
    background: transparent;
    padding: 6px 12px;
    font-size: 14px;
    margin: 0 !important;
    border-radius: 8px 8px 0 0;
    cursor: pointer;
    transition: background 0.15s;
}
div[data-testid="stRadio"] > div[role="radiogroup"] > label:hover {
    background: rgba(74,108,247,0.08);
}
div[data-testid="stRadio"] > div[role="radiogroup"] > label:has(input:checked) {
    background: rgba(74,108,247,0.15);
    border-bottom: 3px solid #4a6cf7;
    font-weight: 700;
}
div[data-testid="stRadio"] > div[role="radiogroup"] > label > div:first-child {
    display: none !important; 
}
div[data-testid="stRadio"] svg { 
    display: none !important; 
}
</style>
""", unsafe_allow_html=True)

tab_aktif = st.radio(
    "Navigasi Tab", TAB_LABELS,
    horizontal=True, label_visibility="collapsed", key="tab_navigasi_utama"
)

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 1: RADAR PDRB
# ═══════════════════════════════════════════════════════════════════════════════
if tab_aktif == TAB_LABELS[0]:
    with st.expander("📖 Panduan Penggunaan Tab Radar Berita", expanded=False):
        st.markdown("""
        **Fungsi Tab Ini:** Tempat Anda memantau dan memburu berita fenomena ekonomi secara otomatis dari internet.
        1. **⚙️ Atur Dahulu di Sidebar:** Pastikan **Rentang Waktu Pencarian** (Dari Tanggal & Sampai Tanggal) dan **Pengaturan AI Radar** (Skor Minimum Lolos, toggle Proses Ulang Artikel Lama, toggle Scan Semua Level Wilayah) di sidebar sudah sesuai kebutuhan — pengaturan ini menentukan periode berita yang dicari dan seberapa ketat AI menyaring hasilnya.
        2. **Jalankan Radar:** Pilih salah satu kategori, lalu klik tombol **▶ SCAN**. Mesin akan mencari berita, lalu AI akan menyaring berita yang tidak relevan.
        3. **Status Kategori PDRB:** Menampilkan ringkasan kategori mana yang sudah punya berita (Aman) dan mana yang kosong (Buntu).
        4. **Hasil Scan Radar Berita:** Setelah scan selesai, berita yang lolos akan muncul di sini. Klik **🚀 Kirim ke Ekstraktor** untuk membedah artikel tersebut di Tab 2.
        """)

    # ── [BARU] CHANGE 1: Kriteria & Sistem Skoring ────────────────────────────
    with st.expander("📋 Kriteria Pencarian & Sistem Skoring AI Radar", expanded=False):
        st.markdown("##### 🌍 Strategi Pencarian: 5 Level Fallback Wilayah")
        st.markdown("""
        Radar mencari berita secara bertahap dari wilayah paling spesifik ke umum.
        Jika di level atas tidak ditemukan, AI otomatis turun ke level berikutnya:
        """)
        kolom_level = st.columns(5)
        level_data = [
            ("1", "🏙️", "Kota Magelang", "#003366", "#cce0ff"),
            ("2", "🏘️", "Kab. Magelang", "#0055a5", "#d6e8ff"),
            ("3", "🗺️", "Eks-Kedu (Sekitar Magelang)", "#1a73e8", "#e3f0ff"),
            ("4", "🌏", "Jawa Tengah",   "#2e86de", "#eef4ff"),
            ("5", "🇮🇩", "Nasional",      "#54a0ff", "#f3f8ff"),
        ]
        for col, (no, ikon, nama, warna_txt, warna_bg) in zip(kolom_level, level_data):
            col.markdown(
                f"""<div style="text-align:center; padding:14px 8px; border-radius:10px;
                background:{warna_bg}; border:1px solid {warna_txt}30;">
                <div style="font-size:1.6rem;">{ikon}</div>
                <div style="font-size:0.7rem; color:{warna_txt}; font-weight:800;">LEVEL {no}</div>
                <div style="font-size:0.82rem; font-weight:600; color:#333;">{nama}</div></div>""",
                unsafe_allow_html=True
            )

        st.markdown("##### 🗺️ Aturan Geografi (Lolos / Tolak)")
        st.caption("📌 Fokus tetap Kota Magelang — wilayah lain hanya dianggap relevan sebagai konteks/dampak ke Kota Magelang, bukan topik yang berdiri sendiri.")
        col_lolos, col_tolak = st.columns(2)
        with col_lolos:
            for teks in [
                "✅ Menyebut 'Kota Magelang' secara eksplisit",
                "✅ Berita Kabupaten Magelang/sekitarnya yang JUGA menyebut atau jelas berkaitan dengan Kota Magelang",
                "✅ Membahas kondisi 'Jawa Tengah' / 'Jateng'",
                "✅ Berita NASIONAL dari Kementerian / Badan / Pemerintah Pusat (Kementan, Bulog, BPS Pusat, dll.)",
            ]:
                st.markdown(f'<div class="geo-lolos">{teks}</div>', unsafe_allow_html=True)
        with col_tolak:
            for teks in [
                "❌ Berita Kabupaten Magelang murni (topik administratif/lokal semata) tanpa kaitan ke Kota Magelang",
                "❌ Artikel dari PROVINSI LAIN (Jatim, Bali, Sumsel, dll.) tanpa menyebut Magelang/Jateng",
                "❌ Artikel opini / lifestyle / hiburan tanpa data statistik apapun",
                "❌ Artikel yang SAMA SEKALI tidak berkaitan dengan kategori PDRB yang dicari",
            ]:
                st.markdown(f'<div class="geo-tolak">{teks}</div>', unsafe_allow_html=True)

        st.markdown("##### 🧠 Sistem Skoring AI (Skala 1–10)")
        st.caption(f"Saat ini skor minimum lolos diatur ke **{min_skor}/10** (bisa diubah di sidebar).")
        skor_data = [
            ("10", "🟢", "Ada data angka SPESIFIK + perbandingan waktu + menyebut Kota Magelang langsung", "green"),
            ("9",  "🟢", "Ada data angka spesifik + perbandingan waktu + konteks Jawa Tengah", "green"),
            ("8",  "🟢", "Ada data angka spesifik + relevan kategori + wilayah valid", "green"),
            ("7",  "🟡", "Ada data angka + relevan kategori + wilayah valid, perbandingan waktu kurang eksplisit", "yellow"),
            ("6",  "🟡", "Ada data angka atau pernyataan resmi + relevan + wilayah valid, data kurang spesifik", "yellow"),
            ("5",  "🔴", "Relevan kategori + wilayah valid, tapi minim data konkret", "red"),
            ("3–4","🔴", "Ada kaitan dengan kategori tapi data sangat kurang atau wilayah kurang relevan", "red"),
            ("1–2","🔴", "Artikel tidak relevan kategori, atau berasal dari provinsi lain / Kabupaten Magelang murni yang tidak berdampak ke Kota", "red"),
        ]
        for skor, ikon, keterangan, warna in skor_data:
            css_class = f"skor-row-{warna}"
            st.markdown(
                f'<div class="{css_class}"><b>{ikon} Skor {skor}</b> &nbsp;—&nbsp; {keterangan}</div>',
                unsafe_allow_html=True
            )

        st.markdown("##### ✅ Kriteria Data Fenomena (minimal salah satu)")
        c1, c2, c3 = st.columns(3)
        c1.info("**A. Data Angka Spesifik**\nHarga (Rp), persentase (%), berat (ton/kg), luas (ha), jumlah unit/orang")
        c2.info("**B. Perbandingan Waktu**\nEksplisit ('naik X% dari bulan lalu', y-on-y, q-to-q) maupun implisit/naratif (misal dibandingkan kejadian serupa tahun lalu, proyeksi ke depan)")
        c3.info("**C. Pernyataan Data Resmi**\nKutipan pejabat/instansi pemerintah tentang kondisi sektor")

    st.markdown("## 📡 Radar Pencari Berita Fenomena")

    # ══════════════════════════════════════════════════════════════════════
    # BAGIAN A: JALANKAN RADAR — diisolasi dalam st.fragment
    # ══════════════════════════════════════════════════════════════════════
    @st.fragment
    def _blok_jalankan_radar():
        with st.container(border=True):
            st.markdown("#### 🚀 Jalankan Radar")
            col_kat, col_btn = st.columns([4, 1])
            with col_kat:
                pilihan_kategori = st.selectbox(
                    "Target Kategori:",
                    ["✨ SEMUA KATEGORI (BATCH SCAN)"] + DAFTAR_KATEGORI,
                    label_visibility="collapsed",
                    help="Pilih 1 kategori untuk discan cepat, atau pilih BATCH SCAN untuk memproses semua kategori sekaligus.",
                    key="selectbox_pilihan_kategori"
                )
            with col_btn:
                btn_scan = st.button("▶ SCAN", type="primary", width='stretch',
                                     help="Mulai pencarian dan filter AI.", key="btn_scan_radar")

            if btn_scan:
                if tanggal_mulai > tanggal_selesai:
                    st.error("Perbaiki rentang tanggal di sidebar terlebih dahulu.")
                else:
                    mulai_str   = tanggal_mulai.strftime("%Y-%m-%d")
                    selesai_str = tanggal_selesai.strftime("%Y-%m-%d")
                    try:
                        _hitung_triwulan(mulai_str)
                    except ValueError as e:
                        st.error(f"⚠️ {e}")
                        st.stop()

                    if pilihan_kategori == "✨ SEMUA KATEGORI (BATCH SCAN)":
                        prog   = st.progress(0)
                        status = st.empty()
                        def cb_progress(kat, idx, total):
                            prog.progress(idx / total)
                            status.info(f"🔄 [{idx}/{total}] Memindai: **{kat}**...")
                        with st.spinner("Memulai Batch Scan — mohon tunggu, Anda bisa minum kopi dulu ☕..."):
                            hasil_batch = batch_scan_semua_kategori(
                                DAFTAR_KATEGORI, mulai_str, selesai_str,
                                min_skor=min_skor, paksa_proses_ulang=paksa_ulang,
                                scan_semua_level=scan_semua,
                                callback_progress=cb_progress
                            )
                        prog.empty(); status.empty()
                        r = hasil_batch["ringkasan"]
                        st.success(
                            f"✅ Batch Scan Selesai! Berita ditemukan di **{r['sukses']} kategori** "
                            f"({r['persen_sukses']}%). Silakan cek tabel antrean di bawah."
                        )
                        time.sleep(2)
                        st.rerun()
                    else:
                        hasil = {}
                        with st.status(
                            f"📡 Radar memindai: **{pilihan_kategori}**...", expanded=True
                        ) as status_box:
                            log_box = st.container(height=280, border=True)
                            log_placeholder = log_box.empty()
                            log_lines = []
                            def cb_log(pesan: str):
                                log_lines.append(pesan)
                                log_placeholder.code("\n".join(log_lines[-30:]), language=None)
                            hasil = scan_kategori(
                                pilihan_kategori, mulai_str, selesai_str,
                                min_skor=min_skor,
                                paksa_proses_ulang=paksa_ulang,
                                scan_semua_level=True,
                                aktifkan_fallback=True,
                                callback_log=cb_log,
                            )
                            if hasil["status"] == "sukses":
                                status_box.update(
                                    label=f"✅ Ditemukan {hasil['jumlah_valid']} artikel valid!",
                                    state="complete"
                                )
                            else:
                                status_box.update(
                                    label="❌ Tidak ada berita ditemukan.",
                                    state="error"
                                )

                        if hasil["status"] == "sukses":
                            st.success(
                                f"✅ Ditemukan **{hasil['jumlah_valid']} artikel** valid! "
                                f"Mengarahkan ke antrean..."
                            )
                            st.session_state.kategori_terpilih_antrean = pilihan_kategori
                        else:
                            st.error(hasil.get("pesan_utama", "Tidak ada berita ditemukan."))
                            with st.expander("💡 Saran Keyword Manual dari AI"):
                                for kw in hasil.get("saran_keyword", []):
                                    st.markdown(f"- `{kw}`")
                                st.markdown("**Coba cari manual di sumber berikut:**")
                                for s in hasil.get("saran_sumber", []):
                                    st.markdown(f"- [{s}](https://{s})")
                        time.sleep(1.5)
                        st.rerun()

    _blok_jalankan_radar()

    st.markdown("---")

    # ── BAGIAN B: DASHBOARD STATUS ────────────────────────────────────────────
    st.markdown("#### 📊 Status Kategori PDRB")
    semua_status = ambil_semua_status_kategori(triwulan_berjalan)

    if not semua_status:
        st.info("ℹ️ Belum ada data untuk triwulan ini. Jalankan Radar terlebih dahulu.")
    else:
        df = pd.DataFrame(semua_status)
        ada    = df[df["jumlah_artikel_valid"] > 0]
        kosong = df[df["jumlah_artikel_valid"] == 0]

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("📋 Total Dipindai",  len(df),
                  help="Jumlah total kategori yang sudah pernah di-scan pada triwulan ini.")
        c2.metric("🟢 Ada Berita",       len(ada),
                  help="Kategori yang sudah memiliki minimal 1 artikel di database.")
        c3.metric("🔴 Kosong/Buntu",     len(kosong),
                  help="Kategori yang belum ditemukan beritanya sama sekali.")
        c4.metric("📈 Coverage",
                  f"{round(len(ada)/len(df)*100)}%" if len(df) else "0%",
                  help="Persentase kelengkapan fenomena BPS untuk triwulan ini.")

        col_ada, col_buntu = st.columns(2)
        with col_ada:
            st.markdown("**✅ Kategori dengan Berita**")
            if not ada.empty:
                st.dataframe(
                    ada[["kategori_pdrb", "jumlah_artikel_valid", "terakhir_scan"]]
                    .rename(columns={"kategori_pdrb": "Kategori",
                                     "jumlah_artikel_valid": "Artikel",
                                     "terakhir_scan": "Terakhir Scan"}),
                    width='stretch', hide_index=True
                )
        with col_buntu:
            st.markdown("**⚠️ Kategori Butuh Perhatian**")
            if not kosong.empty:
                st.dataframe(
                    kosong[["kategori_pdrb", "terakhir_scan"]]
                    .rename(columns={"kategori_pdrb": "Kategori",
                                     "terakhir_scan": "Terakhir Scan"}),
                    width='stretch', hide_index=True
                )

    st.markdown("---")

    # ══════════════════════════════════════════════════════════════════════
    # BAGIAN C: ANTREAN ARTIKEL (Hasil Scan Radar Berita) — diisolasi dalam st.fragment
    # ══════════════════════════════════════════════════════════════════════
    @st.fragment
    def _blok_hasil_scan_radar():
        st.markdown(
            "#### 📥 Hasil Scan Radar Berita",
            help="Daftar berita hasil Radar yang lolos seleksi dan siap dibedah oleh AI Ekstraktor."
        )

        col_filter, col_input_manual = st.columns([2, 2])
        with col_filter:
            opsi_antrean   = ["— Pilih Kategori —"] + DAFTAR_KATEGORI
            index_terpilih = 0
            if st.session_state.kategori_terpilih_antrean in opsi_antrean:
                index_terpilih = opsi_antrean.index(st.session_state.kategori_terpilih_antrean)
            kat_antrean = st.selectbox(
                "Tampilkan hasil scan dari kategori:", opsi_antrean, index=index_terpilih,
                help="Pilih kategori untuk melihat berita hasil radar yang tertangkap.",
                key="selectbox_kat_antrean"
            )
            st.session_state.kategori_terpilih_antrean = kat_antrean

        with col_input_manual:
            st.markdown("**📎 Atau input URL manual (dari Google atau lainnya):**")
            url_manual = st.text_input("URL Berita Manual:", placeholder="https://...",
                                       label_visibility="collapsed", key="text_url_manual")
            if (st.button("📤 Kirim ke Ekstraktor Tab 2", width='stretch',
                          help="Melewati scan radar dan langsung mengirim link ke meja Ekstraktor.",
                          key="btn_kirim_manual")
                    and url_manual):
                st.session_state.target_url = url_manual
                st.toast("URL berhasil dikirim! Silakan buka Tab 2 (Ekstraktor Fenomena).", icon="✅")

        if kat_antrean != "— Pilih Kategori —":
            artikel_db = ambil_artikel_valid(kat_antrean, triwulan_berjalan, min_skor=min_skor)
            if not artikel_db:
                st.info("🎉 Kosong! Semua artikel di kategori ini sudah diekstrak atau belum ada scan baru.")
            else:
                st.markdown(f"**{len(artikel_db)} artikel menunggu ekstraksi:**")
                for art in artikel_db:
                    skor  = art["skor_relevansi"]
                    badge = "badge-skor-hijau" if skor >= 8 else "badge-skor-kuning"
                    label = "🟢" if skor >= 8 else "🟡"
                    with st.container():
                        st.markdown(f"""
                        <div class="kartu-artikel">
                            <b>{art['judul_berita']}</b><br>
                            <a href="{art['url_berita']}" target="_blank"><small>🔗 {art['url_berita']}</small></a><br><br>
                            <span class="{badge}">{label} Skor AI: {skor}/10</span>
                            &nbsp; {'✅ Ada Angka' if art['ada_data_angka'] else '❌ Tanpa Angka'}
                            &nbsp; {'✅ Ada Perbandingan' if art['ada_perbandingan'] else '❌ Tanpa Perbandingan'}
                            <div class="alasan-box">💬 {art['alasan_ai']}</div>
                        </div>
                        """, unsafe_allow_html=True)
                        ca, cb, cc = st.columns([2, 2, 8])
                        with ca:
                            if st.button("🚀 Ekstrak Berita Ini", key=f"eks_{art['id']}",
                                         type="primary", help="Membawa artikel ini ke Tab 2."):
                                st.session_state.target_url = art["url_berita"]
                                st.toast("Berita dikirim! Silakan buka Tab Ekstraktor.", icon="🚀")
                        with cb:
                            if st.button("❌ Tolak (Hapus)", key=f"tolak_{art['id']}",
                                         help="Buang artikel ini dari antrean."):
                                tandai_artikel_ditolak(art["url_berita"])
                                st.toast("Artikel dibuang ke tempat sampah.", icon="🗑️")
                                time.sleep(0.5)
                                st.rerun(scope="fragment")

    _blok_hasil_scan_radar()


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 2: EKSTRAKTOR FENOMENA
# ═══════════════════════════════════════════════════════════════════════════════
elif tab_aktif == TAB_LABELS[1]:
    with st.expander("📖 Panduan Penggunaan Tab Ekstraktor Fenomena", expanded=False):
        st.markdown("""
        **Fungsi Tab Ini:** Meja operasi AI untuk membedah artikel.
        1. **Mulai Ekstrak:** Tekan tombol untuk memerintahkan AI membaca artikel dan memecahnya menjadi 12 poin penting.
        2. **Human-in-the-Loop:** AI tidak selalu 100% sempurna. Anda **WAJIB** mengecek dan mengedit isi kotak jawaban jika ada yang kurang tepat.
        3. **Finalisasi:** Menekan tombol 'Finalisasi' akan menghapus artikel dari antrean Radar (Tab 1) **dan otomatis menyimpannya ke database** untuk download massal di Tab History.
        """)

    st.markdown("## 📝 Meja Ekstraksi Fenomena BPS")

    with st.container(border=True):
        st.markdown("#### 1️⃣ Masukkan URL Berita")
        url_input = st.text_input(
            "URL Berita:",
            value=st.session_state.target_url,
            placeholder="Tempel link berita di sini...",
            label_visibility="collapsed",
            key="input_url_ekstraktor"
        )
        btn_ekstrak = st.button("🤖 MULAI EKSTRAK", type="primary",
                                width='stretch',
                                help="AI akan mulai membaca artikel secara penuh.",
                                key="btn_mulai_ekstrak")

    if btn_ekstrak:
        if not url_input.strip():
            st.error("URL tidak boleh kosong!")
        elif not any(KEYS.values()):
            st.error("Tidak ada API Key yang terisi! Tambahkan di file `.env`.")
        else:
            st.session_state.ekstraksi_url_aktif = url_input.strip()
            st.session_state.hasil_ekstraksi     = None
            st.session_state.json_final_siap     = None
            with st.status("⚙️ Menjalankan Mesin Ekstraksi Lapis 7...", expanded=True) as status_box:
                st.write("🕵️‍♂️ 1/2. Membaca situs web via Bypass Scraper...")
                hasil_scrape = scrape_berita(url_input.strip())
                if hasil_scrape["status"] == "error":
                    status_box.update(label="Scraping Gagal!", state="error")
                    st.error(f"Pesan: {hasil_scrape['pesan']}")
                    st.stop()
                else:
                    st.session_state.ekstraksi_url_aktif = hasil_scrape.get(
                        "url", st.session_state.ekstraksi_url_aktif
                    )
                    st.write(
                        f"✅ Web terbaca via **{hasil_scrape['metode']}** "
                        f"({len(hasil_scrape['teks'])} karakter)."
                    )
                    st.write("🧠 2/2. AI Menganalisis 12 Variabel BPS... (Tunggu 10–20 Detik)")
                    hasil_ai = ekstrak_fenomena_ai(KEYS, hasil_scrape)

                    if hasil_ai["status"] == "error":
                        status_box.update(label="AI Tumbang / Limit Kuota!", state="error")
                        st.error(f"Pesan: {hasil_ai['pesan']}")
                        st.stop()
                    else:
                        model_pakai = hasil_ai["data"].get("_model_digunakan", "AI")
                        st.write(f"✅ Otak AI sukses memecah data menggunakan **{model_pakai}**.")
                        st.session_state.hasil_ekstraksi = hasil_ai["data"]
                        status_box.update(label="🎉 Ekstraksi Sukses!", state="complete")

    st.markdown("---")
    if st.session_state.hasil_ekstraksi is not None:
        data = st.session_state.hasil_ekstraksi

        st.markdown("#### 2️⃣ Validasi & Edit Hasil AI")
        st.info(
            f"🧠 AI yang bertugas: **{data.get('_model_digunakan', 'AI')}** — "
            "Ingat, AI bisa salah baca (Halusinasi). Silakan periksa dan ketik ulang jika ada yang kurang tepat."
        )

        with st.form("form_finalisasi"):
            col1, col2 = st.columns(2)
            with col1:
                tema       = st.text_input("1. Tema Topik",            value=_ke_str(data.get("tema_topik", "")))
                judul_tgl  = st.text_input("2. Judul & Tanggal Terbit", value=_ke_str(data.get("judul_dan_tanggal", "")))
                sumber     = st.text_input("3. Sumber & Link Media",    value=_ke_str(data.get("sumber_dan_link", "")))
                lokasi     = st.text_input("7. Lokasi Spesifik",        value=_ke_str(data.get("lokasi_spesifik", "")))
                periode    = st.text_input("9. Periode Kejadian",       value=_ke_str(data.get("periode_kejadian", "")))
                kata_kunci = st.text_input("10. Kata Kunci / Hashtag",  value=_ke_str(data.get("kata_kunci", "")))
            with col2:
                angka      = st.text_area("5. Data Angka Kuantitatif",
                                          value=_ke_str(data.get("data_angka", "")), height=120)
                intervensi = st.text_area("8. Intervensi Pemerintah",
                                          value=_ke_str(data.get("intervensi_pemerintah", "")), height=120)
                sentimen   = st.selectbox(
                    "11. Sentimen Dampak", ["Positif", "Negatif", "Netral"],
                    index=(["Positif", "Negatif", "Netral"].index(data.get("sentimen_dampak", "Netral"))
                           if data.get("sentimen_dampak") in ["Positif", "Negatif", "Netral"] else 2)
                )
                perbandingan = st.selectbox(
                    "12. Jenis Perbandingan",
                    ["y-on-y", "q-to-q", "harga", "Tidak ada informasi"],
                    index=(["y-on-y", "q-to-q", "harga", "Tidak ada informasi"].index(
                               data.get("kategori_perbandingan", "Tidak ada informasi"))
                           if data.get("kategori_perbandingan") in
                              ["y-on-y", "q-to-q", "harga", "Tidak ada informasi"] else 3)
                )
            ringkasan = st.text_area("4. Ringkasan Fenomena (4-5 Kalimat)",
                                     value=_ke_str(data.get("ringkasan_fenomena", "")), height=160)
            kutipan   = st.text_area("6. Kutipan Tokoh & Narasumber",
                                     value=_ke_str(data.get("kutipan_tokoh", "")), height=130)
            st.markdown("---")
            submit = st.form_submit_button(
                "✅ FINALISASI & TANDAI SELESAI", type="primary",
                width='stretch',
                help="Menyimpan hasil ke database agar laporan bisa didownload.",
                key="submit_finalisasi"
            )

            if submit:
                model_info = data.get("_model_digunakan", "")
                tandai_artikel_diekstrak(st.session_state.ekstraksi_url_aktif)
                st.session_state.json_final_siap = {
                    "tema_topik"           : tema,
                    "judul_dan_tanggal"    : judul_tgl,
                    "sumber_dan_link"      : sumber,
                    "ringkasan_fenomena"   : ringkasan,
                    "data_angka"           : angka,
                    "kutipan_tokoh"        : kutipan,
                    "lokasi_spesifik"      : lokasi,
                    "intervensi_pemerintah": intervensi,
                    "periode_kejadian"     : periode,
                    "kata_kunci"           : kata_kunci,
                    "sentimen_dampak"      : sentimen,
                    "kategori_perbandingan": perbandingan,
                    "_url_sumber"          : st.session_state.ekstraksi_url_aktif,
                    "_waktu_ekstraksi"     : datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "_model_digunakan"     : model_info,
                }
                st.session_state.target_url      = ""
                st.session_state.hasil_ekstraksi = None

                simpan_hasil_ekstraksi(
                    url=st.session_state.ekstraksi_url_aktif,
                    json_final=st.session_state.json_final_siap
                )

                # ── [BARU] CHANGE 2: Pop-up & banner notifikasi ───────────
                st.toast("Hasil ekstraksi berhasil disimpan ke database!", icon="✅")
                st.success(
                    "🎉 **Berhasil difinalisasi!** Hasil ekstraksi ini sudah tersimpan secara otomatis.\n\n"
                    "📦 Temukan di tab **🗄️ HISTORY BERITA** → bagian *'Download Semua Hasil Ekstraksi (Massal)'* "
                    "— bisa diunduh sekaligus dalam format **Excel, CSV, atau JSON**.\n\n"
                    "⬇️ Untuk mengunduh artikel ini saja, gunakan tombol download di bawah."
                )

    # ── DOWNLOAD BUTTONS — DI LUAR FORM ──
    if st.session_state.json_final_siap:
        jf = st.session_state.json_final_siap
        st.markdown("#### 📤 Unduh Hasil Ekstraksi Ini (Pilih Format)")

        LABEL_MAP_DL = [
            ("tema_topik",            "Tema Topik"),
            ("judul_dan_tanggal",     "Judul & Tanggal"),
            ("sumber_dan_link",       "Sumber & Link"),
            ("ringkasan_fenomena",    "Ringkasan Fenomena"),
            ("data_angka",            "Data Angka"),
            ("kutipan_tokoh",         "Kutipan Tokoh"),
            ("lokasi_spesifik",       "Lokasi Spesifik"),
            ("intervensi_pemerintah", "Intervensi Pemerintah"),
            ("periode_kejadian",      "Periode Kejadian"),
            ("kata_kunci",            "Kata Kunci"),
            ("sentimen_dampak",       "Sentimen Dampak"),
            ("kategori_perbandingan", "Kategori Perbandingan"),
        ]

        col_dl1, col_dl2, col_dl3 = st.columns(3)

        with col_dl1:
            st.download_button(
                "⬇️ Download Excel (.xlsx)",
                data=_buat_excel_ekstraksi(jf),
                file_name=_nama_file("EkstraksiFenomena", "xlsx"),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width='stretch', type="primary", key="dl_xlsx"
            )
            st.caption("📊 Excel berformat rapi — paling direkomendasikan untuk laporan BPS")

        with col_dl2:
            df_csv  = pd.DataFrame([{"Variabel": lbl, "Nilai": jf.get(k, "")} for k, lbl in LABEL_MAP_DL])
            csv_buf = io.StringIO()
            df_csv.to_csv(csv_buf, index=False)
            st.download_button(
                "⬇️ Download CSV",
                data=csv_buf.getvalue(),
                file_name=_nama_file("EkstraksiFenomena", "csv"),
                mime="text/csv", width='stretch', key="dl_csv"
            )
            st.caption("📄 CSV — mudah digabungkan di database tabular")

        with col_dl3:
            st.download_button(
                "⬇️ Download JSON",
                data=json.dumps(jf, indent=4, ensure_ascii=False).encode("utf-8"),
                file_name=_nama_file("EkstraksiFenomena", "json"),
                mime="application/json", width='stretch', key="dl_json"
            )
            st.caption("🗂️ JSON — untuk integrasi antar aplikasi/developer")

        with st.expander("👁️ Preview Hasil Akhir Cetak", expanded=False):
            df_preview = pd.DataFrame([
                {"No": i+1, "Variabel BPS": lbl, "Hasil Ekstraksi": jf.get(k, "")}
                for i, (k, lbl) in enumerate(LABEL_MAP_DL)
            ])
            st.dataframe(df_preview, width='stretch', hide_index=True,
                         column_config={"Hasil Ekstraksi": st.column_config.TextColumn(width="large")})

        if st.button("🔄 Mulai Kerjakan Artikel Baru", key="btn_artikel_baru"):
            st.session_state.json_final_siap = None
            st.rerun()


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 3: HISTORY BERITA
# ═══════════════════════════════════════════════════════════════════════════════
elif tab_aktif == TAB_LABELS[2]:
    with st.expander("📖 Panduan Penggunaan Tab History Berita", expanded=False):
        st.markdown("""
        **Fungsi Tab Ini:** Pusat arsip dan ekspor data SI-PENA.
        1. **Tabel Riwayat Radar:** Semua artikel yang pernah ditemukan & disaring Radar, lengkap dengan skor AI dan statusnya.
        2. **Tabel Semua Hasil Ekstraksi:** Semua artikel yang sudah difinalisasi di Tab 2, berisi 12 variabel lengkap. Download massal tersedia dalam Excel, CSV, dan JSON.
        """)

    st.markdown("## 🗄️ Tabel Riwayat Pencarian Radar")

    conn = None
    try:
        conn = get_connection()
        df_riwayat = pd.read_sql_query("""
            SELECT
                judul_berita      AS "Judul Berita",
                kategori_pdrb     AS "Kategori PDRB",
                triwulan          AS "Triwulan",
                skor_relevansi    AS "Skor AI",
                status            AS "Status",
                tanggal_ditemukan AS "Ditemukan",
                tanggal_diekstrak AS "Diekstrak",
                url_berita        AS "URL"
            FROM riwayat_artikel
            ORDER BY tanggal_ditemukan DESC
        """, conn)
    except Exception:
        df_riwayat = pd.DataFrame()
    finally:
        if conn is not None:
            conn.close()

    if df_riwayat.empty:
        st.info("Belum ada riwayat artikel di database. Jalankan Radar terlebih dahulu.")
    else:
        col_f1, col_f2, col_f3 = st.columns(3)
        with col_f1:
            filter_status = st.multiselect(
                "Filter Status:",
                ["ditemukan", "diekstrak", "tidak_lolos", "ditolak_user"],
                default=["ditemukan", "diekstrak"],
                key="multiselect_filter_status"
            )
        with col_f2:
            filter_tw = st.multiselect(
                "Filter Triwulan:",
                df_riwayat["Triwulan"].unique().tolist() if not df_riwayat.empty else [],
                default=df_riwayat["Triwulan"].unique().tolist()[:1] if not df_riwayat.empty else [],
                key="multiselect_filter_tw"
            )
        with col_f3:
            filter_kat = st.multiselect(
                "Filter Kategori (Opsional):",
                df_riwayat["Kategori PDRB"].unique().tolist() if not df_riwayat.empty else [],
                key="multiselect_filter_kat"
            )

        df_tampil = df_riwayat.copy()
        if filter_status: df_tampil = df_tampil[df_tampil["Status"].isin(filter_status)]
        if filter_tw:     df_tampil = df_tampil[df_tampil["Triwulan"].isin(filter_tw)]
        if filter_kat:    df_tampil = df_tampil[df_tampil["Kategori PDRB"].isin(filter_kat)]

        st.markdown(f"**Menampilkan {len(df_tampil)} dari {len(df_riwayat)} total entri di Database**")
        st.dataframe(df_tampil.drop(columns=["URL"]), width='stretch', hide_index=True)

        # ── [CHANGE 4 + 5]: Excel profesional + nama file standar ──────────
        st.markdown("**📤 Export Tabel Riwayat Radar:**")
        col_e1, col_e2, col_e3 = st.columns(3)

        with col_e1:
            # ── [CHANGE 4]: Excel berformat rapi (bukan plain) ──
            st.download_button(
                "⬇️ Download Excel (.xlsx)",
                data=_buat_excel_riwayat(df_tampil.drop(columns=["URL"])),
                file_name=_nama_file("RiwayatRadar", "xlsx"),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width='stretch', type="primary", key="dl_riwayat_xlsx"
            )
        with col_e2:
            csv_exp = df_tampil.to_csv(index=False).encode("utf-8")
            st.download_button(
                "⬇️ Download CSV", data=csv_exp,
                file_name=_nama_file("RiwayatRadar", "csv"),
                mime="text/csv", width='stretch', key="dl_riwayat_csv"
            )
        with col_e3:
            json_riwayat = df_tampil.to_json(orient="records", force_ascii=False, indent=2)
            st.download_button(
                "⬇️ Download JSON", data=json_riwayat.encode("utf-8"),
                file_name=_nama_file("RiwayatRadar", "json"),
                mime="application/json", width='stretch', key="dl_riwayat_json"
            )

    # ── DOWNLOAD MASSAL HASIL EKSTRAKSI ──────────────────────────────────────
    st.markdown("---")
    st.markdown("## 📦 Download Semua Hasil Ekstraksi (Massal)")
    st.caption("Seluruh berita yang sudah pernah difinalisasi di Tab 2, dalam 1 tabel lengkap.")

    conn = None
    try:
        conn = get_connection()
        df_ekstraksi = pd.read_sql_query("""
            SELECT
                waktu_ekstraksi       AS "Waktu Ekstraksi",
                tema_topik            AS "Tema/Topik",
                judul_dan_tanggal     AS "Judul & Tanggal",
                sumber_dan_link       AS "Sumber & Link",
                ringkasan_fenomena    AS "Ringkasan Fenomena",
                data_angka            AS "Data Angka",
                kutipan_tokoh         AS "Kutipan Tokoh",
                lokasi_spesifik       AS "Lokasi Spesifik",
                intervensi_pemerintah AS "Intervensi Pemerintah",
                periode_kejadian      AS "Periode Kejadian",
                kata_kunci            AS "Kata Kunci",
                sentimen_dampak       AS "Sentimen",
                kategori_perbandingan AS "Jenis Perbandingan",
                model_ai              AS "Model AI",
                url_berita            AS "URL"
            FROM hasil_ekstraksi
            ORDER BY waktu_ekstraksi DESC
        """, conn)
    except Exception:
        df_ekstraksi = pd.DataFrame()
    finally:
        if conn is not None:
            conn.close()

    if df_ekstraksi.empty:
        st.info("Belum ada hasil ekstraksi yang disimpan. Finalisasi berita di Tab 2 terlebih dahulu.")
    else:
        st.markdown(f"**Total {len(df_ekstraksi)} hasil ekstraksi tersimpan.**")
        st.dataframe(df_ekstraksi.drop(columns=["URL"]), width='stretch', hide_index=True)

        # ── [CHANGE 3 + 4 + 5]: 3 tombol download + Excel profesional + nama standar ──
        col_d1, col_d2, col_d3 = st.columns(3)
        with col_d1:
            # ── [CHANGE 4]: Excel berformat rapi ──
            st.download_button(
                "⬇️ Download Excel (.xlsx)",
                data=_buat_excel_semua_ekstraksi(df_ekstraksi.drop(columns=["URL"])),
                file_name=_nama_file("SemuaEkstraksi", "xlsx"),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width='stretch', type="primary", key="dl_eks_xlsx"
            )
            st.caption("📊 Excel berformat rapi — direkomendasikan untuk laporan BPS")
        with col_d2:
            csv_all = df_ekstraksi.to_csv(index=False).encode("utf-8")
            st.download_button(
                "⬇️ Download CSV",
                data=csv_all,
                file_name=_nama_file("SemuaEkstraksi", "csv"),
                mime="text/csv", width='stretch', key="dl_eks_csv"
            )
            st.caption("📄 CSV — mudah digabungkan di database tabular")
        with col_d3:
            # ── [CHANGE 3]: Tambah tombol JSON untuk Semua Ekstraksi ──
            json_eks = df_ekstraksi.to_json(orient="records", force_ascii=False, indent=2)
            st.download_button(
                "⬇️ Download JSON",
                data=json_eks.encode("utf-8"),
                file_name=_nama_file("SemuaEkstraksi", "json"),
                mime="application/json", width='stretch', key="dl_eks_json"
            )
            st.caption("🗂️ JSON — untuk integrasi antar aplikasi/developer")


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 4: DASHBOARD ANALISIS
# ═══════════════════════════════════════════════════════════════════════════════
elif tab_aktif == TAB_LABELS[3]:
    st.markdown("## 📈 Dashboard Analisis Fenomena")
    st.caption("Visualisasi interaktif untuk memantau tren berita ekonomi dan performa mesin SI-PENA.")

    conn = None
    try:
        conn = get_connection()
        df_dash = pd.read_sql_query("SELECT * FROM riwayat_artikel", conn)
    except Exception:
        df_dash = pd.DataFrame()
    finally:
        if conn is not None:
            conn.close()

    if df_dash.empty:
        st.info("ℹ️ Belum ada data untuk divisualisasikan. Silakan jalankan Radar terlebih dahulu.")
    else:
        df_dash['Tanggal Ditemukan'] = pd.to_datetime(df_dash['tanggal_ditemukan']).dt.date

        col_m1, col_m2, col_m3, col_m4 = st.columns(4)
        col_m1.metric("📰 Total Berita Tersimpan", len(df_dash),
                      help="Total seluruh berita yang pernah ditangkap Radar.")
        col_m2.metric("🏷️ Kategori Terdampak",     df_dash['kategori_pdrb'].nunique(),
                      help="Jumlah sektor ekonomi (PDRB) yang memiliki minimal 1 berita.")
        col_m3.metric("🧠 Rata-rata Skor AI",       round(df_dash['skor_relevansi'].mean(), 2),
                      help="Skor rata-rata kualitas berita yang disaring AI (Skala 1-10).")
        col_m4.metric("✅ Berita Diekstrak",         len(df_dash[df_dash['status'] == 'diekstrak']),
                      help="Jumlah berita yang sudah selesai diolah di Meja Ekstraktor (Tab 2).")

        st.markdown("---")

        col_c1, col_c2 = st.columns([6, 4])
        with col_c1:
            st.markdown("**📉 Tren Penemuan Berita Harian**")
            st.caption("Melihat lonjakan (spike) kemunculan berita pada hari-hari tertentu.")
            tren_harian = df_dash.groupby('Tanggal Ditemukan').size().reset_index(name='Jumlah Berita')
            chart_tren  = alt.Chart(tren_harian).mark_line(
                point=True, color='#4a6cf7', strokeWidth=3
            ).encode(
                x=alt.X('Tanggal Ditemukan:T', title='Tanggal'),
                y=alt.Y('Jumlah Berita:Q',     title='Jumlah Berita'),
                tooltip=['Tanggal Ditemukan', 'Jumlah Berita']
            ).interactive().properties(height=300)
            st.altair_chart(chart_tren, width='stretch')

        with col_c2:
            st.markdown("**📊 Status Antrean Berita**")
            st.caption("Proporsi berita yang 'ngantre' vs yang sudah diselesaikan.")
            status_dist = df_dash['status'].value_counts().reset_index()
            status_dist.columns = ['Status', 'Jumlah']
            chart_status = alt.Chart(status_dist).mark_arc(innerRadius=65).encode(
                theta=alt.Theta(field="Jumlah", type="quantitative"),
                color=alt.Color(field="Status", type="nominal",
                                scale=alt.Scale(scheme='category10')),
                tooltip=['Status', 'Jumlah']
            ).properties(height=300)
            st.altair_chart(chart_status, width='stretch')

        st.markdown("---")
        st.markdown("**🏆 Top 10 Kategori PDRB Paling Banyak Diberitakan**")
        st.caption("Sektor ekonomi mana yang sedang menjadi sorotan / tren terpanas di media.")
        top_kat = df_dash['kategori_pdrb'].value_counts().head(10).reset_index()
        top_kat.columns = ['Kategori', 'Jumlah Berita']
        chart_bar = alt.Chart(top_kat).mark_bar(color='#28a745', cornerRadiusEnd=4).encode(
            x=alt.X('Jumlah Berita:Q', title='Total Berita Ditemukan'),
            y=alt.Y('Kategori:N', sort='-x', title=''),
            tooltip=['Kategori', 'Jumlah Berita']
        ).properties(height=350)
        st.altair_chart(chart_bar, width='stretch')


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 5: KELOLA KEYWORD
# ═══════════════════════════════════════════════════════════════════════════════
elif tab_aktif == TAB_LABELS[4]:
    st.markdown("## ⚙️ Manajemen Keyword Pencarian")
    st.caption("Tambah, ubah, atau hapus keyword untuk setiap kategori PDRB. Perubahan langsung aktif tanpa restart.")

    kw_data = _load_keywords()
    kat_edit = st.selectbox("Pilih Kategori untuk Diedit:", list(kw_data.keys()), key="selectbox_kat_edit")

    if kat_edit:
        col_m, col_j, col_n = st.columns(3)
        level_map = {
            "🏙️ Level Kota Magelang": ("magelang", col_m),
            "🌏 Level Jawa Tengah":   ("jateng",   col_j),
            "🇮🇩 Level Nasional":     ("nasional", col_n),
        }
        new_keywords = {}
        for label, (key, col) in level_map.items():
            with col:
                st.markdown(f"**{label}**")
                existing = "\n".join(kw_data[kat_edit].get(key, []))
                slug_kat = re.sub(r'[^a-z0-9]+', '_', kat_edit.lower()).strip('_')
                edited   = st.text_area(
                    "Keyword (satu per baris):",
                    value=existing, height=200,
                    key=f"kw_{slug_kat}_{key}",
                    help="Satu keyword per baris. Hapus baris untuk menghapus keyword."
                )
                new_keywords[key] = [k.strip() for k in edited.split("\n") if k.strip()]

        if st.button("💾 Simpan Perubahan", type="primary", width='stretch', key="btn_simpan_keyword"):
            kw_data[kat_edit] = new_keywords
            _save_keywords(kw_data)
            st.success(f"✅ Keyword untuk '{kat_edit}' berhasil disimpan!")
            st.rerun()

    st.markdown("---")
    st.markdown("**➕ Tambah Kategori Baru**")
    nama_baru = st.text_input("Nama Kategori Baru:", key="text_nama_baru")
    if st.button("Tambah Kategori", key="btn_tambah_kategori") and nama_baru:
        kw_data = _load_keywords()
        if nama_baru not in kw_data:
            kw_data[nama_baru] = {"magelang": [], "jateng": [], "nasional": []}
            _save_keywords(kw_data)
            st.success(f"✅ Kategori '{nama_baru}' ditambahkan! Pilih dari dropdown di atas untuk mengisi keyword-nya.")
            st.rerun()
        else:
            st.warning("Kategori sudah ada.")


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 6: TENTANG SI-PENA
# ═══════════════════════════════════════════════════════════════════════════════
elif tab_aktif == TAB_LABELS[5]:

    # ── Hero Banner ───────────────────────────────────────────────────────────
    st.markdown("""
    <div class="hero-banner">
        <h1>✒️ SI-PENA</h1>
        <div class="tagline">Sistem Informasi Pencari Berita &amp; Ekstraksi Fenomena Ekonomi</div>
        <div class="desc">
            Platform berbasis AI yang dirancang khusus untuk membantu analis BPS Kota Magelang
            menemukan, menyaring, dan mengekstrak fenomena ekonomi dari ribuan artikel berita online
            secara otomatis — cepat, terstruktur, dan siap laporan.
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Latar Belakang ────────────────────────────────────────────────────────
    st.markdown("### 📌 Latar Belakang")
    col_lb1, col_lb2 = st.columns([3, 2])
    with col_lb1:
        st.markdown("""
        Dalam proses penyusunan data **PDRB (Produk Domestik Regional Bruto)** Kota Magelang,
        analis BPS membutuhkan referensi fenomena ekonomi yang terjadi setiap triwulan —
        mulai dari fluktuasi harga komoditas, kondisi sektor industri, hingga kebijakan pemerintah
        yang berdampak pada perekonomian lokal.

        Selama ini, proses pencarian dilakukan **secara manual**: membuka search engine,
        menelusuri satu per satu artikel, menentukan relevansi, menyalin data penting, dan memformatnya
        ke dalam formulir. Proses ini memakan waktu berjam-jam untuk setiap kategori PDRB.

        **SI-PENA hadir sebagai solusi**: mengotomasi seluruh alur kerja tersebut dengan bantuan
        kecerdasan buatan (AI), sehingga analis dapat fokus pada validasi dan analisis — bukan
        pencarian manual.
        """)
    with col_lb2:
        st.markdown("""
        <div style="background: linear-gradient(135deg, #f0f7ff, #e8f4ff);
                    border-radius: 12px; padding: 24px; text-align: center;">
            <div style="font-size: 2.5rem; font-weight: 900; color: #003366;">51</div>
            <div style="font-size: 0.85rem; color: #555; margin-bottom: 20px;">Kategori PDRB yang didukung</div>
            <div style="font-size: 2.5rem; font-weight: 900; color: #003366;">5</div>
            <div style="font-size: 0.85rem; color: #555; margin-bottom: 20px;">Level fallback wilayah</div>
            <div style="font-size: 2.5rem; font-weight: 900; color: #003366;">9</div>
            <div style="font-size: 0.85rem; color: #555; margin-bottom: 20px;">Model AI dengan auto-fallback</div>
            <div style="font-size: 2.5rem; font-weight: 900; color: #003366;">12</div>
            <div style="font-size: 0.85rem; color: #555;">Variabel BPS yang diekstrak otomatis</div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("---")

    # ── Fitur Utama ───────────────────────────────────────────────────────────
    st.markdown("### 🚀 Fitur Utama")
    fitur_baris1 = st.columns(3)
    fitur_baris2 = st.columns(3)

    fitur = [
        ("📡", "Radar Pencari Berita",
         "Mencari berita secara otomatis dari DuckDuckGo News & Google News RSS. "
         "Dilengkapi 5 level fallback wilayah (Kota → Kab → Kedu → Jateng → Nasional) "
         "dan sistem scoring AI 1–10 untuk menyaring berita yang benar-benar relevan."),
        ("📝", "Ekstraktor Fenomena AI",
         "Membaca artikel secara penuh dan mengekstrak 12 variabel BPS secara otomatis: "
         "tema, ringkasan fenomena, data angka kuantitatif, kutipan tokoh, lokasi spesifik, "
         "intervensi pemerintah, periode kejadian, sentimen, dan lainnya."),
        ("🗄️", "History & Download Massal",
         "Menyimpan semua riwayat pencarian dan hasil ekstraksi ke database SQLite lokal. "
         "Ekspor data lengkap dalam format Excel (berformat rapi), CSV, dan JSON — "
         "bisa diunduh satu per satu maupun sekaligus massal."),
        ("📈", "Dashboard Analisis",
         "Visualisasi interaktif yang menampilkan tren penemuan berita harian, "
         "distribusi status artikel, dan top 10 kategori PDRB yang paling banyak diberitakan — "
         "untuk memantau performa mesin dan tren ekonomi."),
        ("⚙️", "Manajemen Keyword",
         "Edit, tambah, atau hapus keyword pencarian langsung dari antarmuka Streamlit — "
         "tanpa perlu menyentuh kode program. Perubahan tersimpan ke file JSON dan "
         "langsung aktif di sesi scan berikutnya."),
        ("🤖", "Multi-AI Auto-Fallback",
         "Sistem load balancing dengan 9 model AI (Groq, Gemini, Mistral, Cerebras) "
         "dan dukungan multiple API key per provider. Jika satu model limit/error, "
         "sistem otomatis beralih ke model cadangan berikutnya."),
    ]

    for i, (ikon, judul, deskripsi) in enumerate(fitur):
        kolom = fitur_baris1[i] if i < 3 else fitur_baris2[i - 3]
        kolom.markdown(f"""
        <div class="fitur-card">
            <div class="icon">{ikon}</div>
            <h4>{judul}</h4>
            <p>{deskripsi}</p>
        </div>
        """, unsafe_allow_html=True)
        kolom.markdown("")   # sedikit spasi

    st.markdown("---")

    # ── Alur Kerja ────────────────────────────────────────────────────────────
    st.markdown("### 🔄 Alur Kerja Pipeline Radar")
    st.caption("Setiap kali tombol SCAN ditekan, sistem menjalankan 6 tahapan berurutan secara otomatis:")

    alur = [
        ("🔑", "Query Expansion (Modul A)",
         "Nama kategori PDRB (misal 'Tanaman Pangan') diterjemahkan menjadi keyword "
         "pencarian jurnalistik natural menggunakan kamus statis atau AI fallback. "
         "Menghasilkan keyword untuk 3 level wilayah: Magelang, Jateng, Nasional."),
        ("🔍", "Multi-Source Search (Modul B)",
         "Keyword dicari di 3 mesin sekaligus: DuckDuckGo News, Google News RSS, "
         "dan DuckDuckGo Web (fallback). URL Google News di-resolve ke URL artikel asli "
         "secara otomatis. Hasil digabung dan deduplikasi."),
        ("🗃️", "Database Filter (Modul C)",
         "Setiap URL dicek ke database SQLite. URL yang sudah pernah diekstrak dilewati "
         "agar tidak terjadi pekerjaan ganda. URL baru/belum diproses dilanjutkan ke tahap scraping."),
        ("📥", "Parallel Scraping (Modul D)",
         "Semua URL baru di-scrape secara paralel menggunakan ThreadPoolExecutor. "
         "3 metode scraping berlapis: Jina Reader API → Direct Request → Wayback Machine. "
         "Teks hasil scraping dibersihkan dari noise navigasi dan elemen web."),
        ("🤖", "AI Pre-Screening (Modul E)",
         "Setiap artikel dibaca oleh AI dan diberi skor relevansi 1–10 berdasarkan "
         "kriteria geografi, kelengkapan data angka, dan kesesuaian kategori. "
         "Artikel dengan skor di bawah threshold dibuang, sisanya masuk antrean."),
        ("📝", "Ekstraksi 12 Variabel (Modul F)",
         "Staf memilih artikel dari antrean, lalu AI membaca penuh dan mengekstrak "
         "12 variabel BPS. Staf memvalidasi dan mengedit hasil AI, kemudian "
         "menfinalisasi — data tersimpan ke database dan siap diunduh."),
    ]

    for ikon, judul, deskripsi in alur:
        st.markdown(f"""
        <div class="alur-step">
            <div class="nomor">{alur.index((ikon, judul, deskripsi)) + 1}</div>
            <div class="konten">
                <h5>{ikon} {judul}</h5>
                <p>{deskripsi}</p>
            </div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("---")

    # ── Teknologi ─────────────────────────────────────────────────────────────
    st.markdown("### 🛠️ Teknologi yang Digunakan")
    col_t1, col_t2, col_t3, col_t4 = st.columns(4)

    with col_t1:
        st.markdown("**🖥️ Framework & Backend**")
        for tech, warna in [
            ("Streamlit", "#ff4b4b"), ("Python 3.10+", "#3776ab"),
            ("SQLite", "#003b57"), ("Pandas", "#150458"),
        ]:
            st.markdown(
                f'<span class="tech-chip" style="background:{warna}20; color:{warna}; border:1px solid {warna}40;">'
                f'{tech}</span>', unsafe_allow_html=True
            )

    with col_t2:
        st.markdown("**🤖 Model AI**")
        for tech, warna in [
            ("Groq — GPT-OSS 120B", "#f55036"), ("Gemini 3.1 Flash-Lite", "#4285f4"),
            ("Gemini 3.5 Flash", "#0f9d58"), ("Gemma 4 26B/31B", "#fbbc05"),
            ("Cerebras (GPT-OSS/GLM/Gemma)", "#7c3aed"), ("Mistral Small", "#ff7000"),
        ]:
            st.markdown(
                f'<span class="tech-chip" style="background:{warna}20; color:{warna}; border:1px solid {warna}40;">'
                f'{tech}</span>', unsafe_allow_html=True
            )

    with col_t3:
        st.markdown("**🔍 Sumber Berita**")
        for tech, warna in [
            ("DuckDuckGo News", "#de5833"), ("Google News RSS", "#4285f4"),
            ("DuckDuckGo Web", "#de5833"), ("Jina Reader API", "#0d6efd"),
            ("Wayback Machine", "#795548"),
        ]:
            st.markdown(
                f'<span class="tech-chip" style="background:{warna}20; color:{warna}; border:1px solid {warna}40;">'
                f'{tech}</span>', unsafe_allow_html=True
            )

    with col_t4:
        st.markdown("**📦 Library Utama**")
        for tech, warna in [
            ("openpyxl", "#217346"), ("feedparser", "#ff6b35"),
            ("requests + ddgs", "#0d6efd"), ("Altair (Visualisasi)", "#f9a03c"),
        ]:
            st.markdown(
                f'<span class="tech-chip" style="background:{warna}20; color:{warna}; border:1px solid {warna}40;">'
                f'{tech}</span>', unsafe_allow_html=True
            )

    st.markdown("---")

    # ── Tim & Kontak ──────────────────────────────────────────────────────────
    st.markdown("### 👥 Pengembang & Kontak")
    col_tim1, col_tim2 = st.columns(2)

    with col_tim1:
        st.markdown("""
        <div style="border: 1px solid rgba(0,51,102,0.2); border-radius: 12px;
                    padding: 24px; background: rgba(0,51,102,0.03);">
            <div style="font-size: 1.8rem; margin-bottom: 8px;">🏛️</div>
            <div style="font-weight: 800; font-size: 1.05rem; color: #003366;">
                BPS Kota Magelang
            </div>
            <div style="font-size: 0.88rem; color: #555; margin-top: 6px;">
                Badan Pusat Statistik — Kota Magelang<br>
                Jl. Jend. Gatot Soebroto No. 54-D, Jurangombo Selatan,<br>
                Kec. Magelang Selatan, Kota Magelang, Jawa Tengah 56123
            </div>
            <div style="margin-top: 14px; font-size: 0.85rem;">
                🌐 <a href="https://magelangkota.bps.go.id" target="_blank"
                     style="color:#003366;">magelangkota.bps.go.id</a>
            </div>
        </div>
        """, unsafe_allow_html=True)

    with col_tim2:
        st.markdown("""
        <div style="border: 1px solid rgba(74,108,247,0.25); border-radius: 12px;
                    padding: 24px; background: rgba(74,108,247,0.03);">
            <div style="font-size: 1.8rem; margin-bottom: 8px;">💡</div>
            <div style="font-weight: 800; font-size: 1.05rem; color: #003366;">
                Catatan Pengembangan
            </div>
            <div style="font-size: 0.88rem; color: #555; margin-top: 6px; line-height: 1.7;">
                SI-PENA dikembangkan sebagai bagian dari program Magang mahasiswa Fakultas
                Ilmu Komputer Universitas Brawijaya di BPS Kota Magelang.<br><br>
                Platform ini bersifat <em>open-source internal</em> dan dapat dikembangkan lebih lanjut.
            </div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("")
    st.caption(
        "SI-PENA v1.0  •  Made with ❤️  •  "
        f"BPS Kota Magelang {datetime.now().year}"
    )